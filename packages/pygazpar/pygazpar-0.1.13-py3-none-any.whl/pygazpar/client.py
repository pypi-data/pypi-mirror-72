import os
import time
import glob
from selenium import webdriver
from datetime import datetime
from openpyxl import load_workbook
from pygazpar.enum import PropertyNameEnum

HOME_URL = 'https://monespace.grdf.fr'
LOGIN_URL = HOME_URL + '/monespace/connexion'
WELCOME_URL = HOME_URL + '/monespace/particulier/accueil'
DATA_URL = HOME_URL + '/monespace/particulier/consommation/consommations'
DATA_FILENAME = 'Consommations gaz_*.xlsx'

DEFAULT_TMP_DIRECTORY = '/tmp'
DEFAULT_FIREFOX_WEBDRIVER = os.getcwd() + '/geckodriver'
DEFAULT_WAIT_TIME = 30

class LoginError(Exception):
    """ Client has failed to login in GrDF Web site (check username/password)"""
    pass

class Client(object):
    def __init__(self, username, password, firefox_webdriver_executable = DEFAULT_FIREFOX_WEBDRIVER, wait_time = DEFAULT_WAIT_TIME, tmp_directory = DEFAULT_TMP_DIRECTORY):
        self.__username = username
        self.__password = password
        self.__firefox_webdriver_executable = firefox_webdriver_executable
        self.__wait_time = wait_time
        self.__tmp_directory = tmp_directory
        self.__data = []

    def data(self):
        return self.__data

    def update(self):

        # XLSX is in the TMP directory
        data_file_path_pattern = self.__tmp_directory + '/' + DATA_FILENAME

        # We remove an eventual existing file (from a previous run that has not deleted it)
        file_list = glob.glob(data_file_path_pattern)
        for filename in file_list:
            if os.path.isfile(filename):
                os.remove(filename)

        # Initialize the Firefox WebDriver
        profile = webdriver.FirefoxProfile()
        options = webdriver.FirefoxOptions()
        options.headless = True
        profile.set_preference('browser.download.folderList', 2)  # custom location
        profile.set_preference('browser.download.manager.showWhenStarting', False)
        profile.set_preference('browser.helperApps.alwaysAsk.force', False)
        profile.set_preference('browser.download.dir', self.__tmp_directory)
        profile.set_preference('browser.helperApps.neverAsk.saveToDisk', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        driver = webdriver.Firefox(executable_path=self.__firefox_webdriver_executable, firefox_profile=profile, options=options, service_log_path=self.__tmp_directory + '/geckodriver.log')
        try:
            driver.set_window_position(0, 0)
            driver.set_window_size(1200, 1200)

            driver.implicitly_wait(self.__wait_time)
            
            ## Login URL
            driver.get(LOGIN_URL)
            
            # Fill login form
            email_element = driver.find_element_by_id("_EspacePerso_WAR_EPportlet_:seConnecterForm:email")
            password_element = driver.find_element_by_id("_EspacePerso_WAR_EPportlet_:seConnecterForm:passwordSecretSeConnecter")
            
            email_element.send_keys(self.__username)
            password_element.send_keys(self.__password)
            
            submit_button_element = driver.find_element_by_id('_EspacePerso_WAR_EPportlet_:seConnecterForm:meConnecter')
            submit_button_element.click()
            
            # Eventually, close Advertisement Popup Windows
            try:
                close_popup_element = driver.find_elements_by_css_selector('.abtasty-modal__close > svg')
                close_popup_element[0].click()
            except:
                # Do nothing, because the Pop up may not appear.
                pass

            # Once we find the 'Accéder' button from the main page, we are logged on successfully.
            try:
                driver.find_element_by_xpath("//div[2]/div[2]/div/a/div")
            except:
                # Perhaps, login has failed.
                if driver.current_url == WELCOME_URL:
                    # We're good.
                    pass
                elif driver.current_url == LOGIN_URL:
                    raise LoginError("GrDF sign in has failed, please check your username/password")
                else:
                    raise

            # Page 'votre consommation'
            driver.get(DATA_URL)          

            # Wait for the data page to load completely.
            time.sleep(self.__wait_time)

            # Select daily consumption
            daily_consumption_element = driver.find_element_by_xpath("//table[@id='_eConsoconsoDetaille_WAR_eConsoportlet_:idFormConsoDetaille:panelTypeGranularite1']/tbody/tr/td[3]/label")
            daily_consumption_element.click()

            # Download file
            download_button_element = driver.find_element_by_xpath("//button[@onclick=\"envoieGATelechargerConsoDetaille('particulier', 'jour_kwh');\"]/span")
            download_button_element.click()
            
            # Timestamp of the data.
            data_timestamp = datetime.now().isoformat()

            # Wait a few for the download to complete
            time.sleep(self.__wait_time)
            
            # Load the XLSX file into the data structure
            file_list = glob.glob(data_file_path_pattern)

            for filename in file_list:
                wb = load_workbook(filename = filename)
                ws = wb['Historique par jour']
                for rownum in range(8, len(ws['B'])+1):
                    row = {}
                    if ws.cell(column=2, row=rownum).value != None:
                        row[PropertyNameEnum.DATE.value] = ws.cell(column=2, row=rownum).value                        
                        row[PropertyNameEnum.START_INDEX_M3.value] = ws.cell(column=3, row=rownum).value
                        row[PropertyNameEnum.END_INDEX_M3.value] = ws.cell(column=4, row=rownum).value
                        row[PropertyNameEnum.VOLUME_M3.value] = ws.cell(column=5, row=rownum).value
                        row[PropertyNameEnum.ENERGY_KWH.value] = ws.cell(column=6, row=rownum).value
                        row[PropertyNameEnum.CONVERTER_FACTOR.value] = ws.cell(column=7, row=rownum).value
                        row[PropertyNameEnum.LOCAL_TEMPERATURE.value] = ws.cell(column=8, row=rownum).value
                        row[PropertyNameEnum.TYPE.value] = ws.cell(column=9, row=rownum).value
                        row[PropertyNameEnum.TIMESTAMP.value] = data_timestamp
                        self.__data.append(row)
                wb.close()
            
                os.remove(filename)

        finally:
            # Quit the driver
            driver.quit()
